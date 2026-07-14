"""Google-download ``SpreadsheetStore`` adapter (Drive API + openpyxl).

The read path is DOWNLOAD-centric (the design decision): instead of per-cell Sheets API queries,
it downloads the whole file once (Drive ``files.export`` → xlsx for a native Google Sheet, or
``files.get?alt=media`` for an uploaded ``.xlsx``), parses it locally with openpyxl, and caches
the parsed workbook by ``sheet_id`` with a TTL. Auth is the tenant's OAuth access token (the host
mints/refreshes it and passes it in; the scheduler's DSN analogue). Writes (``swap_rows``) go back
surgically via the Sheets ``values.batchUpdate`` API so formatting/other tabs are untouched.

``pip install cogno-praxis[coordinator]`` (openpyxl + httpx). The two HTTP seams — ``_fetch`` and
``_push_values`` — are the only network methods, so unit tests mock them and exercise the parse /
cache / range / swap-request logic with zero network.
"""

from __future__ import annotations

import io
import time
from typing import Optional

import httpx

_DRIVE = "https://www.googleapis.com/drive/v3/files"
_SHEETS = "https://sheets.googleapis.com/v4/spreadsheets"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DEFAULT_TTL = 600.0                 # 10 min, parent parity


class GoogleSheetsStore:
    """A :class:`~cogno_praxis.coordinator.store.SpreadsheetStore` backed by Google Drive.

    Bound to one tenant's OAuth token at construction (single-tenant-per-instance, like the
    scheduler's Pg adapter). ``now`` is injectable for deterministic cache-TTL tests."""

    def __init__(self, access_token: str, *, ttl_seconds: float = _DEFAULT_TTL,
                 timeout: float = 30.0, now: Optional[object] = None) -> None:
        self._token = access_token
        self._ttl = ttl_seconds
        self._timeout = timeout
        self._now = now or time.monotonic
        # sheet_id → (fetched_at, {tab_lower: grid})
        self._cache: dict[str, tuple[float, dict[str, list[list[str]]]]] = {}

    # ── network seams (mocked in unit tests) ─────────────────────────────────────────
    def _fetch(self, sheet_id: str) -> bytes:
        """Download the spreadsheet as xlsx bytes. Try export (native Google Sheet); on 403
        failedPrecondition (an uploaded .xlsx can't be exported) fall back to raw media."""
        headers = {"Authorization": f"Bearer {self._token}"}
        with httpx.Client(timeout=self._timeout) as c:
            r = c.get(f"{_DRIVE}/{sheet_id}/export", params={"mimeType": _XLSX_MIME},
                      headers=headers)
            if r.status_code == 403:
                r = c.get(f"{_DRIVE}/{sheet_id}", params={"alt": "media"}, headers=headers)
            r.raise_for_status()
            return r.content

    def _push_values(self, sheet_id: str, data: list[dict]) -> None:
        """Sheets values.batchUpdate — ``data`` = ``[{"range": "Tab!A2:C2", "values": [[...]]}]``."""
        headers = {"Authorization": f"Bearer {self._token}"}
        body = {"valueInputOption": "RAW", "data": data}
        with httpx.Client(timeout=self._timeout) as c:
            r = c.post(f"{_SHEETS}/{sheet_id}/values:batchUpdate", headers=headers, json=body)
            r.raise_for_status()

    # ── parse + cache ────────────────────────────────────────────────────────────────
    @staticmethod
    def _parse(xlsx: bytes) -> dict[str, list[list[str]]]:
        """xlsx bytes → ``{tab_lower: [[cell,…],…]}`` (all cells stringified, None→'')."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
        out: dict[str, list[list[str]]] = {}
        for ws in wb.worksheets:
            grid: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                grid.append(["" if v is None else str(v).strip() for v in row])
            out[ws.title.strip().lower()] = grid
        wb.close()
        return out

    def _tabs(self, sheet_id: str) -> dict[str, list[list[str]]]:
        hit = self._cache.get(sheet_id)
        if hit and (self._now() - hit[0]) < self._ttl:
            return hit[1]
        tabs = self._parse(self._fetch(sheet_id))
        self._cache[sheet_id] = (self._now(), tabs)
        return tabs

    @staticmethod
    def _first_row(a1_range: str) -> int:
        import re
        m = re.search(r"[A-Za-z]+(\d+)", a1_range or "")
        return int(m.group(1)) if m else 1

    # ── the port ─────────────────────────────────────────────────────────────────────
    def read_range(self, sheet_id: str, tab: str, a1_range: str) -> list[list[str]]:
        grid = self._tabs(sheet_id).get(tab.strip().lower(), [])
        start = self._first_row(a1_range) - 1
        return [list(r) for r in grid[start:]]

    def swap_rows(self, sheet_id: str, tab: str, a1_range: str, row_a: int, row_b: int,
                  *, content_cols: list[int]) -> None:
        """Swap the ``content_cols`` between two range-relative rows, writing both back via the
        Sheets API (dates/fixed columns stay put). Invalidates the cache for this sheet."""
        grid = self._tabs(sheet_id).get(tab.strip().lower(), [])
        off = self._first_row(a1_range) - 1
        abs_a, abs_b = off + row_a, off + row_b            # 0-based absolute rows in the sheet
        if abs_a >= len(grid) or abs_b >= len(grid):
            raise IndexError(f"swap_rows out of range: {row_a}/{row_b} in {sheet_id}/{tab}")
        row_a_vals, row_b_vals = list(grid[abs_a]), list(grid[abs_b])
        updates: list[dict] = []
        for c in content_cols:
            if c < len(row_a_vals) and c < len(row_b_vals):
                a1_a = f"{_col_letter(c)}{abs_a + 1}"       # Sheets rows/cols are 1-based
                a1_b = f"{_col_letter(c)}{abs_b + 1}"
                updates.append({"range": f"'{tab}'!{a1_a}", "values": [[row_b_vals[c]]]})
                updates.append({"range": f"'{tab}'!{a1_b}", "values": [[row_a_vals[c]]]})
        if updates:
            self._push_values(sheet_id, updates)
            self._cache.pop(sheet_id, None)                # force a fresh read next time


def _col_letter(idx0: int) -> str:
    """0-based column index → spreadsheet letter (0→A, 25→Z, 26→AA)."""
    s = ""
    n = idx0 + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
